"""
Script de importação da planilha CONTROLE DPTO - SECO - R06.xlsx para o Firestore.
Uso: python importar_planilha.py
"""

import openpyxl
import json
import urllib.request
import urllib.parse
import datetime
import sys

# ── CONFIG ──────────────────────────────────────────────────────────────────
PLANILHA = r"C:\Users\henrique.moreira\iCloudDrive\CLAUDE - IA\CONTROLE DPTO - SECO - R06.xlsx"
API_KEY  = "AIzaSyDIKxQjZPjAwZuR9zTolBsY8dyGvGRdJxw"
PROJECT  = "isocs-f91df"
EMAIL    = "henrique.moreira@isoeste.com.br"
SENHA    = "123456"

# ── HELPERS ─────────────────────────────────────────────────────────────────
def post_json(url, payload):
    data = json.dumps(payload).encode()
    req  = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def get_token():
    url  = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={API_KEY}"
    resp = post_json(url, {"email": EMAIL, "password": SENHA, "returnSecureToken": True})
    return resp["idToken"]

def firestore_value(v):
    if v is None:
        return {"nullValue": None}
    if isinstance(v, bool):
        return {"booleanValue": v}
    if isinstance(v, int):
        return {"integerValue": str(v)}
    if isinstance(v, float):
        return {"doubleValue": v}
    if isinstance(v, datetime.datetime):
        return {"timestampValue": v.strftime("%Y-%m-%dT%H:%M:%SZ")}
    if isinstance(v, datetime.date):
        return {"timestampValue": datetime.datetime(v.year, v.month, v.day).strftime("%Y-%m-%dT%H:%M:%SZ")}
    return {"stringValue": str(v)}

def to_fs_doc(fields_dict):
    return {"fields": {k: firestore_value(v) for k, v in fields_dict.items() if v is not None and v != "" and str(v) != "None"}}

def create_doc(token, collection, doc_dict):
    url  = f"https://firestore.googleapis.com/v1/projects/{PROJECT}/databases/(default)/documents/{collection}"
    req  = urllib.request.Request(
        url,
        data=json.dumps(to_fs_doc(doc_dict)).encode(),
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {token}"}
    )
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def val(v):
    """Normaliza valor: string vazia e '#VALUE!' viram None."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "#VALUE!", "#N/A", "#REF!", "#NAME?"):
            return None
    return v

def date_str(v):
    """Converte datetime/date em string ISO ou None."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None

# ── LER PLANILHA ────────────────────────────────────────────────────────────
print("📂 Abrindo planilha...")
wb = openpyxl.load_workbook(PLANILHA, data_only=True)

def read_rows(ws, header_row, data_start):
    headers = [c.value for c in list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]]
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(v is not None for v in row[:5]):
            continue
        rows.append({str(headers[i]) if headers[i] else f"col_{i}": row[i]
                     for i in range(len(headers)) if i < len(row)})
    return rows

orcamentos_raw  = read_rows(wb["OR\u00c7AMENTOS"],  4, 5)
projetos_raw    = read_rows(wb["PROJETOS"],         4, 5)
arquitetura_raw = read_rows(wb["ARQUITETURA"],      4, 5)
art_raw         = read_rows(wb["ART"],              4, 5)

print(f"   Orçamentos : {len(orcamentos_raw)} linhas")
print(f"   Projetos   : {len(projetos_raw)} linhas")
print(f"   Arquitetura: {len(arquitetura_raw)} linhas")
print(f"   ART        : {len(art_raw)} linhas")

# ── AUTENTICAR ──────────────────────────────────────────────────────────────
print("\n🔑 Autenticando no Firebase...")
token = get_token()
print("   OK")

# ── IMPORTAR ORÇAMENTOS ─────────────────────────────────────────────────────
print("\n📤 Importando Orçamentos...")
erros_orc = 0
for i, r in enumerate(orcamentos_raw, 1):
    doc = {
        "status"          : val(r.get("STATUS")),
        "fluig"           : val(r.get("FLUIG")),
        "mes"             : val(r.get("M\ufffdS") or r.get("MÊS") or r.get("M?S")),
        "tag"             : val(r.get("TAG")),
        "reparticao"      : val(r.get("REPARTI\ufffd\ufffdO") or r.get("REPARTIÇÃO") or r.get("REPARTI??O")),
        "tipologia"       : val(r.get("TIPOLOGIA")),
        "edificacao"      : val(r.get("EDIFICA\ufffd\ufffdO") or r.get("EDIFICAÇÃO") or r.get("EDIFICA??O")),
        "aguas"           : val(r.get("\ufffdGUAS") or r.get("ÁGUAS") or r.get("?GUAS")),
        "escopo"          : val(r.get("ESCOPO")),
        "cliente"         : val(r.get("CLIENTE")),
        "projeto"         : val(r.get("PROJETO")),
        "comercial"       : val(r.get("COMERCIAL")),
        "cidade"          : val(r.get("CIDADE")),
        "estado"          : val(r.get("ESTADO")),
        "area"            : val(r.get("\ufffdREA") or r.get("ÁREA") or r.get("?REA")),
        "peso"            : val(r.get("PESO")),
        "mf"              : val(r.get("MF")),
        "santri"          : val(r.get("SANTRI")),
        "metodologia"     : val(r.get("METODOLOGIA")),
        "valorRevestimento": val(r.get("R$ REVEST.")),
        "valorEstrutura"  : val(r.get("R$ ESTRUTURA")),
        "valorMO"         : val(r.get("R$ M.O.")),
        "valorTotal"      : val(r.get("VALOR TOTAL")),
        "entradaFluig"    : date_str(r.get("ENTRADA FLUIG")),
        "saidaControladoria": date_str(r.get("SA\ufffdDA CONTROLADORIA") or r.get("SAÍDA CONTROLADORIA")),
        "responsavel"     : val(r.get("RESPONS\ufffdVEL") or r.get("RESPONSÁVEL") or r.get("RESPONS?VEL")),
        "fechado"         : val(r.get("FECHADO?")),
        "mesFechamento"   : val(r.get("M\ufffdS FECHAMENTO") or r.get("MÊS FECHAMENTO")),
        "valorFechado"    : val(r.get("R$ FECHADO")),
        "importadoDaPlanilha": True,
    }
    # remove None
    doc = {k: v for k, v in doc.items() if v is not None}
    try:
        create_doc(token, "orcamentos", doc)
        if i % 50 == 0:
            print(f"   {i}/{len(orcamentos_raw)}...")
    except Exception as e:
        erros_orc += 1
        if erros_orc <= 3:
            print(f"   ⚠️  Linha {i}: {e}")

print(f"   ✅ {len(orcamentos_raw) - erros_orc} importados, {erros_orc} erros")

# ── IMPORTAR PROJETOS ───────────────────────────────────────────────────────
print("\n📤 Importando Projetos...")
erros_proj = 0
for i, r in enumerate(projetos_raw, 1):
    # Helper para pegar campos com encoding variável
    def g(r, *keys):
        for k in keys:
            if r.get(k) is not None:
                return r.get(k)
        # Tenta match parcial
        for rk in r:
            for k in keys:
                if k.replace("?","") in rk.replace("?","") and rk.replace("?","") in k.replace("?",""):
                    return r[rk]
        return None

    doc = {
        "status"          : val(r.get("STATUS")),
        "mf"              : val(r.get("MF")),
        "fluig"           : val(r.get("FLUIG")),
        "mes"             : val(r.get("M\ufffdS") or r.get("MÊS")),
        "tag"             : val(r.get("TAG")),
        "reparticao"      : val(r.get("REPARTI\ufffd\ufffdO") or r.get("REPARTIÇÃO")),
        "tipologia"       : val(r.get("TIPOLOGIA")),
        "edificacao"      : val(r.get("EDIFICA\ufffd\ufffdO") or r.get("EDIFICAÇÃO")),
        "aguas"           : val(r.get("\ufffdGUAS") or r.get("ÁGUAS")),
        "escopo"          : val(r.get("ESCOPO")),
        "cliente"         : val(r.get("CLIENTE")),
        "projeto"         : val(r.get("PROJETO")),
        "comercial"       : val(r.get("COMERCIAL")),
        "cidade"          : val(r.get("CIDADE")),
        "estado"          : val(r.get("ESTADO")),
        "area"            : val(r.get("\ufffdREA") or r.get("ÁREA")),
        "previsaoImportacao": date_str(r.get("PREVIS\ufffd\ufffdO IMPORTA\ufffd\ufffdO") or r.get("PREVISÃO IMPORTAÇÃO")),
        "entradaPropostaCliente": date_str(r.get("PROPOSTA APROVADA")),
        "entradaAprovCliente": date_str(r.get("ENTRADA APROV. CLIENTE")),
        "saidaAprovCliente": date_str(r.get("SA\ufffdDA APROV. CLIENTE") or r.get("SAÍDA APROV. CLIENTE")),
        "entradaCompatibilizacao": date_str(r.get("ENTRADA COMPATIBILIZA\ufffd\ufffdO") or r.get("ENTRADA COMPATIBILIZAÇÃO")),
        "saidaCompatibilizacao": date_str(r.get("SA\ufffdDA COMPATIBILIZA\ufffd\ufffdO") or r.get("SAÍDA COMPATIBILIZAÇÃO")),
        "entradaCadastro" : date_str(r.get("ENTRADA CADASTRO")),
        "entradaControladoria": date_str(r.get("ENTRADA CONTROLADORIA")),
        "saidaControladoria": date_str(r.get("SA\ufffdDA CONTROLADORIA") or r.get("SAÍDA CONTROLADORIA")),
        "saidaCadastro"   : date_str(r.get("SA\ufffdDA CADASTRO") or r.get("SAÍDA CADASTRO")),
        "entradaImplantacao": date_str(r.get("ENTRADA IMPLANTA\ufffd\ufffdO") or r.get("ENTRADA IMPLANTAÇÃO")),
        "saidaImplantacao": date_str(r.get("SA\ufffdDA IMPLANTA\ufffd\ufffdO") or r.get("SAÍDA IMPLANTAÇÃO")),
        "nrMfImplantada"  : val(r.get("N\ufffd MF IMPLANTADA") or r.get("Nº MF IMPLANTADA")),
        "peso"            : val(r.get("PESO")),
        "orcado"          : val(r.get("OR\ufffd\ufffdADO") or r.get("ORÇADO")),
        "importacao"      : val(r.get("IMPORTA\ufffd\ufffdO") or r.get("IMPORTAÇÃO")),
        "valorTotal"      : val(r.get("R$ TOTAL")),
        "importadoDaPlanilha": True,
    }
    doc = {k: v for k, v in doc.items() if v is not None}
    try:
        create_doc(token, "projetos", doc)
        if i % 30 == 0:
            print(f"   {i}/{len(projetos_raw)}...")
    except Exception as e:
        erros_proj += 1
        if erros_proj <= 3:
            print(f"   ⚠️  Linha {i}: {e}")

print(f"   ✅ {len(projetos_raw) - erros_proj} importados, {erros_proj} erros")

# ── IMPORTAR ARQUITETURA ────────────────────────────────────────────────────
print("\n📤 Importando Arquitetura...")
erros_arq = 0
for i, r in enumerate(arquitetura_raw, 1):
    doc = {
        "id"        : val(r.get("ID")),
        "mf"        : val(r.get("MF")),
        "fluig"     : val(r.get("FLUIG")),
        "tipologia" : val(r.get("TIPOLOGIA")),
        "metodo"    : val(r.get("M\ufffdTODO") or r.get("MÉTODO")),
        "tecnico"   : val(r.get("T\ufffdCNICO") or r.get("TÉCNICO")),
        "cliente"   : val(r.get("CLIENTE")),
        "projeto"   : val(r.get("PROJETO")),
        "comercial" : val(r.get("COMERCIAL")),
        "cidade"    : val(r.get("CIDADE")),
        "estado"    : val(r.get("ESTADO")),
        "maoDeObra" : val(r.get("M\ufffdO DE OBRA") or r.get("MÃO DE OBRA")),
        "area"      : val(r.get("\ufffdREA") or r.get("ÁREA")),
        "escopo"    : val(r.get("ESCOPO")),
        "valorTotal": val(r.get("R$ TOTAL")),
        "valorMaterial": val(r.get("R$ MATERIAL")),
        "valorMO"   : val(r.get("R$ M.O.")),
        "valorFechado": val(r.get("R$ FECHADO")),
        "desconto"  : val(r.get("DESCONTO")),
        "kgAco"     : val(r.get("KG A\ufffd\ufffdO") or r.get("KG AÇO")),
        "indice"    : val(r.get("\ufffdNDICE") or r.get("ÍNDICE")),
        "entrada"   : date_str(r.get("ENTRADA")),
        "saida"     : date_str(r.get("SA\ufffdDA") or r.get("SAÍDA")),
        "sla"       : val(r.get("SLA")),
        "lead"      : val(r.get("LEAD")),
        "mes"       : val(r.get("M\ufffdS") or r.get("MÊS")),
        "importadoDaPlanilha": True,
    }
    doc = {k: v for k, v in doc.items() if v is not None}
    try:
        create_doc(token, "arquitetura", doc)
    except Exception as e:
        erros_arq += 1
print(f"   ✅ {len(arquitetura_raw) - erros_arq} importados, {erros_arq} erros")

# ── IMPORTAR ART ────────────────────────────────────────────────────────────
print("\n📤 Importando ART...")
erros_art = 0
for i, r in enumerate(art_raw, 1):
    doc = {
        "status"        : val(r.get("STATUS")),
        "fluig"         : val(r.get("FLUIG")),
        "obra"          : val(r.get("OBRA")),
        "mes"           : val(r.get("M\ufffdS") or r.get("MÊS")),
        "modalidade"    : val(r.get("MODALIDADE")),
        "reparticao"    : val(r.get("REPARTI\ufffd\ufffdO") or r.get("REPARTIÇÃO")),
        "tipologia"     : val(r.get("TIPOLOGIA")),
        "edificacao"    : val(r.get("EDIFICA\ufffd\ufffdO") or r.get("EDIFICAÇÃO")),
        "aguas"         : val(r.get("\ufffdGUAS") or r.get("ÁGUAS")),
        "escopo"        : val(r.get("ESCOPO")),
        "cliente"       : val(r.get("CLIENTE")),
        "comercial"     : val(r.get("COMERCIAL")),
        "cidade"        : val(r.get("CIDADE")),
        "estado"        : val(r.get("ESTADO")),
        "area"          : val(r.get("\ufffdREA") or r.get("ÁREA")),
        "peso"          : val(r.get("PESO")),
        "mf"            : val(r.get("MF")),
        "nrArt"         : val(r.get("N\ufffd ART") or r.get("Nº ART")),
        "valorMaterial" : val(r.get("VALOR MAT")),
        "valorMO"       : val(r.get("VALOR M.O.")),
        "valorTotal"    : val(r.get("VALOR TOTAL")),
        "entrada"       : date_str(r.get("ENTRADA")),
        "saida"         : date_str(r.get("SA\ufffdDA") or r.get("SAÍDA")),
        "responsavelTecnico": val(r.get("RESPONS\ufffdVEL T\ufffdCNICO") or r.get("RESPONSÁVEL TÉCNICO")),
        "fechado"       : val(r.get("FECHADO?")),
        "importadoDaPlanilha": True,
    }
    doc = {k: v for k, v in doc.items() if v is not None}
    try:
        create_doc(token, "art", doc)
    except Exception as e:
        erros_art += 1
print(f"   ✅ {len(art_raw) - erros_art} importados, {erros_art} erros")

print("\n🎉 Importação concluída!")
